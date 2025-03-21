import openpyxl  
 
workbook = openpyxl.load_workbook('./8457 - Separação.xlsx')
worksheet = workbook.active
 

lista_de_os = []
 
for row in range(2, 51):
    dictionary = {
        'OS': worksheet.cell(row, 3).value,
        'FUNCOS': worksheet.cell(row, 40).value,
        'QTEMCX': worksheet.cell(row, 9).value,
        'CODPROD': worksheet.cell(row, 4).value,
    }
    lista_de_os.append(dictionary)
 
OS_details = {}
 
for item in lista_de_os:
    OS = item['OS']
    qtemcx = item['QTEMCX']
    codprod = item['CODPROD']
 
    if OS in OS_details:
        OS_details[OS]['QTEMCX'] += qtemcx
        OS_details[OS]['CODPROD'].add(codprod)
        OS_details[OS]['QTD_ENDERECOS'] = len(OS_details[OS]['CODPROD'])
    else:
        OS_details[OS] = {
            'OS': OS,
            'FUNCOS': item['FUNCOS'],
            'QTEMCX': qtemcx,
            'CODPROD': {codprod},
            'QTD_ENDERECOS': 1
        }
 
 
OSs_em_detalhe = list(OS_details.values())
 
funcionarios = {}
 
for OS_realizada in OSs_em_detalhe:
    FUNCOS = OS_realizada['FUNCOS']
    OS = OS_realizada['OS']
    QTD_ENDERECOS = OS_realizada['QTD_ENDERECOS']
 
    if FUNCOS in funcionarios:
        funcionarios[FUNCOS]['OS'].append(OS)
        funcionarios[FUNCOS]['QTD_ENDERECOS'] += QTD_ENDERECOS
    else:
        funcionarios[FUNCOS] = {
            'OS': [OS],
            'QTD_ENDERECOS': QTD_ENDERECOS
        }
 
print(funcionarios)